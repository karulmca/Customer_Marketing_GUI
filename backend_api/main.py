"""
FastAPI Backend for Company Data Scraper
Exposes existing Python functionality as REST APIs for React frontend
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, status, Request
import os
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import Optional, List, Dict, Any

import sys
import json
import asyncio

# Add parent directory to path for imports  
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__)), 'database_config'))
from database_config.file_upload_processor import FileUploadProcessor
import pandas as pd
import numpy as np
from datetime import datetime
import threading
import uuid
import logging
import threading as _threading

print(f"Deploying commit: {os.getenv('GIT_COMMIT', 'unknown')}")

# APScheduler for background scheduled jobs
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    from apscheduler.triggers.interval import IntervalTrigger
    APSCHEDULER_AVAILABLE = True
except Exception as _apex:
    APSCHEDULER_AVAILABLE = False
    print(f"⚠️ APScheduler not available: {_apex}")

# Add parent directory to path for imports
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.insert(0, parent_dir)

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import existing functionality  
from auth.user_auth import UserAuthenticator
from database_config.db_utils import get_database_connection, check_database_requirements
from database_config.file_upload_processor import FileUploadProcessor
from database_config.config_loader import get_scheduler_interval, is_single_job_per_user_enabled

# MVC controllers disabled for now to fix import issues
MVC_CONTROLLERS_AVAILABLE = False

app = FastAPI(
    title="Company Data Scraper API",
    description="REST API for file upload and data processing",
    version="1.0.0"
)

# Include MVC routers if available
if MVC_CONTROLLERS_AVAILABLE:
    from backend_api.controllers import company_data_router, file_data_router
    app.include_router(company_data_router)
    app.include_router(file_data_router)
    logger.info("✅ MVC controllers registered successfully")

# CORS middleware for React frontend
# Configure CORS origins based on environment
cors_origins = [
    "http://localhost:3000", 
    "http://localhost:3001", 
    "http://127.0.0.1:3000",
    "http://127.0.0.1:3001"
]

# Add production origins
environment = os.getenv("ENVIRONMENT", "development")
if environment == "production":
    cors_origins.extend([
        "https://company-scraper-frontend-nddn.onrender.com",
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        # Add any additional production domains here
    ])

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "HEAD", "PATCH"],
    allow_headers=[
        "accept",
        "accept-encoding",
        "authorization",
        "content-type",
        "dnt",
        "origin",
        "user-agent",
        "x-csrftoken",
        "x-requested-with",
        "access-control-allow-credentials",
        "access-control-allow-headers",
        "access-control-allow-methods",
        "access-control-allow-origin"
    ],
)

# Log CORS configuration for debugging
logger.info(f"🌐 CORS configured for origins: {cors_origins}")
logger.info(f"🔧 Environment: {environment}")

# Mount static files from React build
static_dir = os.path.join(parent_dir, "frontend", "build")
if os.path.exists(static_dir):
    app.mount("/static", StaticFiles(directory=os.path.join(static_dir, "static")), name="static")

# Security
security = HTTPBearer()

# Health check endpoint for Docker
@app.get("/health")
async def health_check():
    """Health check endpoint for Docker containers and load balancers"""
    try:
        # Check database connection
        conn = get_database_connection()
        if conn:
            conn.close()
            db_status = "healthy"
        else:
            db_status = "unhealthy"
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    return {
        "status": "healthy" if db_status == "healthy" else "unhealthy",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0",
        "database": db_status,
        "scheduler": "active" if APSCHEDULER_AVAILABLE else "unavailable"
    }

@app.get("/scheduler-status")
async def scheduler_status():
    """Check detailed scheduler status and jobs, including configured interval"""
    try:
        from database_config.config_loader import get_scheduler_interval
        interval_minutes = get_scheduler_interval()
        if not scheduler:
            return {"error": "Scheduler not initialized", "scheduler_running": False, "interval_minutes": interval_minutes}

        jobs = scheduler.get_jobs()
        job_info = []
        for job in jobs:
            job_info.append({
                "id": job.id,
                "name": job.name,
                "next_run": job.next_run_time.isoformat() if job.next_run_time else None,
                "trigger": str(job.trigger),
                "interval_minutes": interval_minutes
            })

        return {
            "scheduler_running": scheduler.running,
            "jobs": job_info,
            "total_jobs": len(jobs),
            "scheduler_available": APSCHEDULER_AVAILABLE,
            "interval_minutes": interval_minutes
        }
    except Exception as e:
        return {"error": str(e), "scheduler_running": False}

@app.get("/debug/pending-uploads")
async def debug_pending_uploads():
    """Debug endpoint to check pending uploads"""
    try:
        from database_config.file_upload_processor import FileUploadProcessor
        processor = FileUploadProcessor()
        
        # Get all uploads (not just pending)
        all_uploads_query = """
        SELECT id, file_name, processing_status, upload_date, uploaded_by 
        FROM file_upload 
        ORDER BY upload_date DESC 
        LIMIT 10
        """
        
        from database_config.postgresql_config import PostgreSQLConfig
        db_config = PostgreSQLConfig()
        all_uploads = db_config.query_to_dataframe(all_uploads_query)
        
        # Get pending uploads specifically
        pending_df = processor.get_pending_uploads()
        
        return {
            "total_recent_uploads": len(all_uploads) if not all_uploads.empty else 0,
            "recent_uploads": all_uploads.to_dict('records') if not all_uploads.empty else [],
            "pending_uploads_count": len(pending_df) if pending_df is not None and not pending_df.empty else 0,
            "pending_uploads": pending_df.to_dict('records') if pending_df is not None and not pending_df.empty else []
        }
    except Exception as e:
        return {"error": str(e)}

# Global variables
auth_system = UserAuthenticator()
active_sessions = {}  # Store active user sessions
processing_jobs = {}  # Store background processing jobs

# Scheduler globals
scheduler = None
SCHEDULER_JOB_ID = "process_pending_uploads_job"
scheduler_state = {
    "running": False,
    "job_added": False,
    "last_run": None,
    "last_result": None,
    "last_error": None,
}
_scheduler_lock = _threading.Lock()

def _ensure_scheduler():
    """Create the background scheduler if available and not yet created."""
    global scheduler
    if not APSCHEDULER_AVAILABLE:
        raise HTTPException(status_code=500, detail="APScheduler is not installed on the server")
    if scheduler is None:
        scheduler = BackgroundScheduler(timezone="UTC")
    # Guard to ensure scheduler is started only once
    if not getattr(scheduler, "_started_once", False):
        if not scheduler.running:
            scheduler.start()
        scheduler._started_once = True
    return scheduler

def _process_pending_uploads():
    """Job: process pending uploads in file_upload table (no overlap)."""
    # Prevent overlapping runs at the function level as well
    if getattr(_process_pending_uploads, "_busy", False):
        logger.info("⏳ Skipping run; previous job still in progress")
        return
    setattr(_process_pending_uploads, "_busy", True)
    try:
        scheduler_state["running"] = True
        scheduler_state["last_error"] = None
        logger.info("🚀 Scheduler job started: processing pending file uploads")

        # Use existing FileUploadProcessor from database_config
        processor = FileUploadProcessor()

        # Get pending uploads with single job per user logic
        try:
            if is_single_job_per_user_enabled():
                pending_df = processor.get_pending_uploads_by_user_queue()
                logger.info("🔒 Using single job per user processing")
            else:
                pending_df = processor.get_pending_uploads()
                logger.info("🔓 Using multi-job processing")
        except Exception as e:
            logger.error(f"Failed to fetch pending uploads: {e}")
            scheduler_state["last_error"] = str(e)
            scheduler_state["last_run"] = datetime.now().isoformat()
            scheduler_state["last_result"] = {"success": False, "processed": 0, "error": str(e)}
            return

        if pending_df is None or pending_df.empty:
            logger.info("ℹ️ No eligible jobs found for processing")
            scheduler_state["last_run"] = datetime.now().isoformat()
            scheduler_state["last_result"] = {"success": True, "processed": 0}
            return

        logger.info(f"📋 Found {len(pending_df)} eligible job(s) for processing")
        # Prefer the consolidated automated runner which encapsulates the manual scrapers
        try:
            from backend_api.automated_job.run_automated_jobs import run_once as automated_run_once
            # Limit the runner to the number of pending rows discovered
            result = automated_run_once(limit=len(pending_df))
            success_count = int(result.get('successful', 0))
            failure_count = int(result.get('failed', 0))
            total = int(result.get('processed', success_count + failure_count))
            logger.info(f"📊 Scheduler batch done (automated runner). Success: {success_count}, Failed: {failure_count}, Total: {total}")
            scheduler_state["last_run"] = datetime.now().isoformat()
            scheduler_state["last_result"] = {
                "success": True,
                "processed": total,
                "successful": success_count,
                "failed": failure_count,
            }
        except Exception as e:
            logger.error(f"❌ Automated runner failed: {e}")
            scheduler_state["last_error"] = str(e)
            scheduler_state["last_run"] = datetime.now().isoformat()
            scheduler_state["last_result"] = {"success": False, "processed": 0, "error": str(e)}
    finally:
        setattr(_process_pending_uploads, "_busy", False)
        scheduler_state["running"] = False

def _clean_for_json_serialization(obj):
    """Clean data to ensure it can be JSON serialized without encoding errors"""
    import numpy as np
    if isinstance(obj, dict):
        return {k: _clean_for_json_serialization(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_clean_for_json_serialization(item) for item in obj]
    elif isinstance(obj, bytes):
        try:
            return obj.decode('utf-8')
        except UnicodeDecodeError:
            return obj.decode('utf-8', errors='replace')
    elif isinstance(obj, str):
        try:
            obj.encode('utf-8')
            return obj
        except UnicodeEncodeError:
            return obj.encode('utf-8', errors='replace').decode('utf-8')
    elif isinstance(obj, (np.integer, int)):
        return int(obj)
    elif isinstance(obj, (np.floating, float)):
        return float(obj)
    elif isinstance(obj, (np.bool_, bool)):
        return bool(obj)
    elif hasattr(obj, 'tolist'):
        return obj.tolist()
    elif obj is None:
        return None
    else:
        return str(obj)

# Pydantic models
class LoginRequest(BaseModel):
    username: str
    password: str

class RegisterRequest(BaseModel):
    username: str
    password: str
    email: str

class FileUploadResponse(BaseModel):
    success: bool
    message: str
    file_id: Optional[str] = None
    preview_data: Optional[Dict] = None

class ProcessingStatus(BaseModel):
    job_id: str
    status: str  # "pending", "processing", "completed", "failed"
    progress: int
    message: str
    result: Optional[Dict] = None

class FileUploadResponse(BaseModel):
    success: bool
    file_upload_id: Optional[str] = None
    message: str
    filename: str
    status: str

class UploadedFile(BaseModel):
    id: str
    filename: str
    uploaded_by: str
    upload_date: Optional[str]
    status: str
    rows_count: Optional[int]
    processing_status: Optional[str]
    processed_at: Optional[str]

class UploadedFilesResponse(BaseModel):
    files: List[UploadedFile]
    count: int

# Authentication endpoints
@app.post("/api/auth/login")
async def login(request: LoginRequest, req: Request):
    """Login endpoint"""
    try:
        # Get client IP address
        client_ip = req.client.host if req.client else "unknown"
        
        result = auth_system.authenticate_user(request.username, request.password, client_ip)
        
        if result['success']:
            session_id = str(uuid.uuid4())
            session_data = {
                'user_info': result.get('user', result.get('user_info', {})),
                'session_token': result.get('session_token', session_id),
                'login_time': datetime.now().isoformat()
            }
            
            # Store in memory
            active_sessions[session_id] = session_data
            
            # Also store in database for persistence
            try:
                from database_config.postgresql_config import PostgreSQLConfig
                import psycopg2
                import json
                
                db_config = PostgreSQLConfig()
                connection_params = db_config.get_connection_params()
                connection = psycopg2.connect(**connection_params)
                cursor = connection.cursor()
                
                # Get user ID for the session
                user_info = session_data.get('user_info', {})
                username = user_info.get('username', 'unknown')
                
                # Find user ID
                cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
                user_result = cursor.fetchone()
                user_id = user_result[0] if user_result else None
                
                cursor.execute("""
                    INSERT INTO user_sessions (user_id, session_token, created_at, expires_at)
                    VALUES (%s, %s, NOW(), NOW() + INTERVAL '24 hours')
                    ON CONFLICT (session_token) 
                    DO UPDATE SET 
                        expires_at = EXCLUDED.expires_at,
                        last_accessed = NOW()
                """, (user_id, session_id))
                
                connection.commit()
                cursor.close()
                connection.close()
                
                print(f"✅ Session stored in database: {session_id}")
                
            except Exception as db_error:
                print(f"⚠️ Warning: Could not store session in database: {str(db_error)}")
                # Continue anyway - memory session still works
            
            return {
                "success": True,
                "message": "Login successful",
                "session_id": session_id,
                "user_info": result.get('user', result.get('user_info', {}))
            }
        else:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail=result.get('message', 'Login failed')
            )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Login error: {str(e)}"
        )

@app.post("/api/auth/register")
async def register(request: RegisterRequest):
    """Register new user endpoint"""
    try:
        result = auth_system.register_user(request.username, request.password, request.email)
        if result['success']:
            return {
                "success": True,
                "message": "User registered successfully"
            }
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=result.get('message', 'Registration failed')
            )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Registration error: {str(e)}")
        

@app.post("/api/auth/logout")
async def logout(session_id: str):
    """Logout endpoint"""
    if session_id in active_sessions:
        del active_sessions[session_id]
    
    return {"success": True, "message": "Logged out successfully"}

# User Management Endpoints
@app.get("/api/auth/users")
async def get_users(session_id: str):
    """Get all users - admin only"""
    try:
        # Verify session
        user_info = verify_session(session_id)
        if not user_info:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        # Check if user is admin (optional - you can remove this if you want all users to see user list)
        # if user_info.get('role') != 'admin':
        #     raise HTTPException(status_code=403, detail="Admin access required")
        
        # Get users from database using PostgreSQL connection
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2

        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)

        cursor = connection.cursor()
        # First, add missing columns if they don't exist
        try:
            cursor.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT TRUE")
            cursor.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_superuser BOOLEAN DEFAULT FALSE")
            connection.commit()
        except Exception as e:
            logger.info(f"Columns may already exist: {e}")

        cursor.execute("""
            SELECT id, username, email, role, created_at, 
                   COALESCE(is_active, TRUE) as is_active,
                   COALESCE(is_superuser, FALSE) as is_superuser
            FROM users 
            ORDER BY created_at DESC
        """)
        
        users = []
        for row in cursor.fetchall():
            users.append({
                'id': row[0],
                'username': row[1],
                'email': row[2],
                'role': row[3],
                'created_at': row[4].isoformat() if row[4] else None,
                'is_active': row[5],
                'is_superuser': row[6]
            })
        
        cursor.close()
        connection.close()
        
        return {"success": True, "users": users}
        
    except Exception as e:
        logger.error(f"Error getting users: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get users: {str(e)}")

@app.post("/api/auth/users")
async def create_user(user_data: dict, session_id: str):
    """Create new user - admin only"""
    try:
        # Verify session
        user_info = verify_session(session_id)
        if not user_info:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        # Create user using existing auth system
        from auth.user_auth import UserAuthenticator
        auth = UserAuthenticator()
        
        # Register the new user
        result = auth.register_user(
            username=user_data.get('username'),
            password=user_data.get('password'),
            email=user_data.get('email')
        )
        
        if result['success']:
            # Update role and status fields if specified
            from database_config.postgresql_config import PostgreSQLConfig
            import psycopg2

            db_config = PostgreSQLConfig()
            connection_params = db_config.get_connection_params()
            connection = psycopg2.connect(**connection_params)
            
            cursor = connection.cursor()
            
            # Build update query for additional fields
            update_fields = []
            values = []
            
            if user_data.get('role'):
                update_fields.append("role = %s")
                values.append(user_data.get('role'))
            
            if 'is_active' in user_data:
                update_fields.append("is_active = %s")
                values.append(user_data.get('is_active', True))
            
            if 'is_superuser' in user_data:
                update_fields.append("is_superuser = %s")
                values.append(user_data.get('is_superuser', False))
            
            if update_fields:
                values.append(user_data.get('username'))
                query = f"UPDATE users SET {', '.join(update_fields)} WHERE username = %s"
                cursor.execute(query, values)
                connection.commit()
            
            cursor.close()
            connection.close()
            
            return {"success": True, "message": "User created successfully"}
        else:
            raise HTTPException(status_code=400, detail=result.get('message', 'Failed to create user'))
            
    except Exception as e:
        logger.error(f"Error creating user: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create user: {str(e)}")

@app.put("/api/auth/users")
async def update_user(user_data: dict, session_id: str):
    """Update user information"""
    try:
        # Verify session
        user_info = verify_session(session_id)
        if not user_info:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        # Get connection and update user using PostgreSQL connection
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2

        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        
        cursor = connection.cursor()
        
        # Build update query dynamically
        update_fields = []
        values = []
        
        if 'username' in user_data and user_data['username']:
            update_fields.append("username = %s")
            values.append(user_data['username'])
        
        if 'email' in user_data:
            update_fields.append("email = %s")
            values.append(user_data['email'])
            
        if 'role' in user_data:
            update_fields.append("role = %s")
            values.append(user_data['role'])
        
        if 'is_active' in user_data:
            update_fields.append("is_active = %s")
            values.append(user_data['is_active'])
        
        if 'is_superuser' in user_data:
            update_fields.append("is_superuser = %s")
            values.append(user_data['is_superuser'])
        
        if update_fields:
            values.append(user_data['id'])  # Add ID for WHERE clause
            query = f"UPDATE users SET {', '.join(update_fields)} WHERE id = %s"
            cursor.execute(query, values)
            connection.commit()
        
        cursor.close()
        connection.close()
        
        return {"success": True, "message": "User updated successfully"}
        
    except Exception as e:
        logger.error(f"Error updating user: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update user: {str(e)}")

@app.delete("/api/auth/users/{user_id}")
async def delete_user(user_id: str, session_id: str):
    """Delete user"""
    try:
        # Verify session
        user_info = verify_session(session_id)
        if not user_info:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        # Prevent self-deletion using PostgreSQL connection
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2

        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        
        cursor = connection.cursor()
        
        # Check if trying to delete self
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user_to_delete = cursor.fetchone()
        
        if user_to_delete and user_to_delete[0] == user_info.get('username'):
            cursor.close()
            connection.close()
            raise HTTPException(status_code=400, detail="Cannot delete your own account")
        
        # Delete user
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        if cursor.rowcount == 0:
            cursor.close()
            connection.close()
            raise HTTPException(status_code=404, detail="User not found")
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return {"success": True, "message": "User deleted successfully"}
        
    except Exception as e:
        logger.error(f"Error deleting user: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete user: {str(e)}")

@app.get("/api/debug/sessions")
async def debug_sessions():
    """Debug endpoint to check active sessions"""
    return {
        "active_sessions": list(active_sessions.keys()),
        "session_count": len(active_sessions)
    }

# Helper function to verify session
def verify_session(session_id: str):
    """Verify active session - enhanced with better persistence and error handling"""
    logger.debug(f"Verifying session_id: {session_id}")
    logger.debug(f"Active sessions in memory: {list(active_sessions.keys())}")
    
    if not session_id or session_id.strip() == "":
        logger.warning("Empty session_id provided")
        return None
    
    # First check memory (fast path)
    if session_id in active_sessions:
        session_data = active_sessions[session_id]
        # Update last accessed time in database
        try:
            _update_session_last_accessed(session_id)
        except Exception as e:
            logger.warning(f"Could not update session last accessed: {e}")
        return session_data
    
    # If not in memory, check database (persistent sessions)
    try:
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2
        
        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        cursor = connection.cursor()
        
        # Check if session exists and is valid (extended expiry time)
        cursor.execute("""
            SELECT us.user_id, us.created_at, u.username, u.email, u.role,
                   us.expires_at, us.last_accessed
            FROM user_sessions us
            JOIN users u ON us.user_id = u.id
            WHERE us.session_token = %s 
            AND us.expires_at > NOW()
        """, (session_id,))
        
        result = cursor.fetchone()
        
        if result:
            user_id, created_at, username, email, role, expires_at, last_accessed = result
            
            # Extend session if it's been accessed recently
            cursor.execute("""
                UPDATE user_sessions 
                SET expires_at = NOW() + INTERVAL '24 hours',
                    last_accessed = NOW()
                WHERE session_token = %s
            """, (session_id,))
            connection.commit()
            
            session_data = {
                'user_info': {
                    'id': user_id,
                    'username': username,
                    'email': email,
                    'role': role
                },
                'session_token': session_id,
                'login_time': created_at.isoformat(),
                'expires_at': expires_at.isoformat() if expires_at else None
            }
            
            # Restore session to memory with extended expiry
            active_sessions[session_id] = session_data
            logger.info(f"✅ Session restored from database: {session_id} for user: {username}")
            
            cursor.close()
            connection.close()
            return session_data
        else:
            logger.warning(f"❌ Session not found or expired in database: {session_id}")
            cursor.close()
            connection.close()
            return None
            
    except Exception as e:
        logger.error(f"❌ Error checking database session: {str(e)}")
    
    # Session not found anywhere
    return None

def _update_session_last_accessed(session_id: str):
    """Update session last accessed time"""
    try:
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2
        
        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        cursor = connection.cursor()
        
        cursor.execute("""
            UPDATE user_sessions 
            SET last_accessed = NOW()
            WHERE session_token = %s
        """, (session_id,))
        connection.commit()
        cursor.close()
        connection.close()
        
    except Exception as e:
        logger.debug(f"Session update error: {e}")

# Enhanced job queue management for concurrent users
job_queue_lock = asyncio.Lock()
user_processing_status = {}  # Track processing status per user

# Cleanup function for stale jobs
async def cleanup_stale_jobs():
    """Clean up jobs that have been running for too long (over 30 minutes)"""
    try:
        async with job_queue_lock:
            current_time = datetime.now()
            stale_users = []
            
            for username, job_info in user_processing_status.items():
                if job_info.get('status') == 'processing':
                    started_str = job_info.get('started')
                    if started_str:
                        try:
                            started_time = datetime.fromisoformat(started_str)
                            time_diff = current_time - started_time
                            
                            # If job has been running for more than 30 minutes, mark as stale
                            if time_diff.total_seconds() > 1800:  # 30 minutes
                                stale_users.append(username)
                                logger.warning(f"🧹 Marking stale job for user {username}: {job_info.get('filename', 'unknown')} (running for {time_diff})")
                        except (ValueError, TypeError):
                            # Invalid timestamp, mark as stale
                            stale_users.append(username)
            
            # Clean up stale jobs
            for username in stale_users:
                user_processing_status[username]['status'] = 'timeout'
                user_processing_status[username]['error'] = 'Job timed out after 30 minutes'
                user_processing_status[username]['completed'] = current_time.isoformat()
    
    except Exception as e:
        logger.error(f"Error during stale job cleanup: {str(e)}")

## Remove duplicate cleanup_scheduler and startup_event logic

# File upload endpoints
@app.post("/api/files/upload")
async def upload_file(
    session_id: str,
    file: UploadFile = File(...)
):
    """Upload Excel file for processing"""
    try:
        # Verify session
        session = verify_session(session_id)
        
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only Excel files (.xlsx, .xls) are supported"
            )
        
        # Read file content directly into memory (NO local storage)
        file_id = str(uuid.uuid4())
        content = await file.read()
        
        # Read and preview file directly from memory
        try:
            import io
            df = pd.read_excel(io.BytesIO(content))
            
            # Clean the data for JSON serialization with robust handling
            def clean_value(value):
                """Clean a single value for JSON serialization"""
                if value is None:
                    return None
                elif pd.isna(value):
                    return None
                elif isinstance(value, (int, np.integer)):
                    return int(value)
                elif isinstance(value, (float, np.floating)):
                    if np.isfinite(value):
                        return float(value)
                    else:
                        return None
                elif isinstance(value, str):
                    return str(value)
                elif hasattr(value, 'isoformat'):  # datetime objects
                    return value.isoformat()
                else:
                    return str(value)
            
            # Get sample data with robust cleaning
            sample_records = []
            if len(df) > 0:
                for _, row in df.head(5).iterrows():
                    clean_record = {}
                    for col in df.columns:
                        clean_record[str(col)] = clean_value(row[col])
                    sample_records.append(clean_record)
            
            preview_data = {
                "filename": file.filename,
                "rows": int(len(df)),
                "columns": int(len(df.columns)),
                "column_names": [str(col) for col in df.columns.tolist()],
                "sample_data": sample_records
            }
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Error reading Excel file: {str(e)}"
            )
        
        # Save file to database directly
        db_file_id = None
        try:
            from database_config.postgresql_config import PostgreSQLConfig
            import psycopg2
            import json
            
            db_config = PostgreSQLConfig()
            connection_params = db_config.get_connection_params()
            connection = psycopg2.connect(**connection_params)
            
            cursor = connection.cursor()
            
            # Prepare data for database insertion
            username = session['user_info'].get('username', 'API_User')
            raw_data = {
                "columns": [str(col) for col in df.columns.tolist()],
                "data": sample_records[:100],  # Store first 100 records as sample
                "metadata": {
                    "total_rows": len(df),
                    "total_columns": len(df.columns),
                    "upload_timestamp": datetime.now().isoformat(),
                    "file_extension": os.path.splitext(file.filename)[1].lower()
                }
            }
            
            # Insert file record
            cursor.execute("""
                INSERT INTO file_upload 
                (file_name, file_path, file_size, original_columns, raw_data, 
                 uploaded_by, processing_status, records_count, file_hash)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                file.filename,
                "database_storage",  # No file path, stored in database
                len(content),  # File size from content length
                json.dumps([str(col) for col in df.columns.tolist()]),
                json.dumps(raw_data),
                username,
                'uploaded',
                len(df),
                str(hash(str(content) + str(datetime.now())))  # Hash from content
            ))
            
            db_file_id = cursor.fetchone()[0]
            connection.commit()
            cursor.close()
            connection.close()
            
            print(f"✅ File saved to database with ID: {db_file_id}")
            
        except Exception as db_error:
            print(f"❌ Database save error: {str(db_error)}")
            db_file_id = None
        
        # Store file info for processing (NO local storage)
        processing_jobs[file_id] = {
            "file_content": content,  # Store content in memory instead of file path
            "filename": file.filename,
            "user_info": session['user_info'],
            "session_token": session['session_token'],
            "status": "uploaded",
            "progress": 0,
            "message": f"File uploaded successfully{' and saved to database' if db_file_id else ' (database save failed)'}",
            "db_file_id": db_file_id  # Link to database record
        }
        
        return FileUploadResponse(
            success=True,
            message="File uploaded successfully",
            file_id=file_id,
            preview_data=preview_data
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload failed for user {username if 'username' in locals() else 'unknown'}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")
    

@app.get("/api/files/status/{file_id}")
async def get_processing_status(file_id: str, session_id: str):
    """Get file processing status"""
    try:
        # Verify session
        verify_session(session_id)
        
        if file_id not in processing_jobs:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found"
            )
        
        job = processing_jobs[file_id]
        
        # Clean the result data to ensure UTF-8 compatibility
        result_data = job.get("result")
        if result_data:
            # Convert any bytes or non-UTF-8 strings to safe UTF-8 strings
            result_data = _clean_for_json_serialization(result_data)
        
        # Also clean the message field
        message = _clean_for_json_serialization(job["message"])
        
        return ProcessingStatus(
            job_id=file_id,
            status=job["status"],
            progress=job["progress"],
            message=message,
            result=result_data
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Status check error: {str(e)}"
        )

@app.get("/api/files/download/{file_id}")
async def download_processed_file(file_id: str, session_id: str):
    """Download processed file"""
    from fastapi.responses import FileResponse
    try:
        # Verify session
        verify_session(session_id)
        
        if file_id not in processing_jobs:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found"
            )
        
        job = processing_jobs[file_id]
        
        if job["status"] != "completed":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File processing not completed"
            )
        
        result = job.get("result", {})
        output_content = result.get("output_content")
        
        if not output_content:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Processed file content not found"
            )
        
        # Return file content directly from memory (NO local files)
        from fastapi.responses import Response
        return Response(
            content=output_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={result.get('output_file', 'processed_file.xlsx')}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Download error: {str(e)}"
        )

@app.get("/api/files/download-processed/{file_id}")
async def download_processed_file_with_linkedin(file_id: str, session_id: str):
    """Download processed file with LinkedIn enrichment data in Excel format"""
    try:
        # Verify session
        verify_session(session_id)
        
        # Get database connection
        db_connection = get_database_connection("postgresql")
        if not db_connection:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Database connection failed"
            )
        
        # Ensure database connection is established
        if not db_connection.connect():
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Failed to connect to database"
            )
        
        # Connect to database
        if not db_connection.connect():
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Failed to establish database connection"
            )
        
        # Query processed company data with LinkedIn enrichment
        query = f"""
        SELECT 
            company_name as "Company Name",
            linkedin_url as "LinkedIn_URL",
            company_website as "Website_URL", 
            company_size as "Company_Size",
            industry as "Industry",
            revenue as "Revenue"
        FROM company_data 
        WHERE file_upload_id = '{file_id}' 
        AND processing_status = 'completed'
        ORDER BY company_name
        """
        
        logger.info(f"Executing query for file_id: {file_id}")
        result_df = db_connection.query_to_dataframe(query)
        
        if result_df is None or result_df.empty:
            logger.warning(f"No processed data found for file_id: {file_id}")
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No processed data found for this file"
            )
        
        logger.info(f"Found {len(result_df)} processed records for file_id: {file_id}")
        
        # Create Excel file in memory
        from io import BytesIO
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Processed Company Data"
        
        # Add data to worksheet
        for r in dataframe_to_rows(result_df, index=False, header=True):
            ws.append(r)
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Get original filename for the processed file name (file_upload.file_name)
        try:
            file_query = f"SELECT file_name FROM file_upload WHERE id = '{file_id}'"
            file_result = db_connection.query_to_dataframe(file_query)
            original_filename = "processed_data.xlsx"
            if file_result is not None and not file_result.empty:
                # prefer file_name column from file_upload
                original_name = None
                if 'file_name' in file_result.columns:
                    original_name = file_result.iloc[0]['file_name']
                elif 'original_filename' in file_result.columns:
                    original_name = file_result.iloc[0]['original_filename']

                if original_name and isinstance(original_name, str):
                    name_parts = original_name.rsplit('.', 1)
                    original_filename = f"processed_{name_parts[0]}.xlsx"
                else:
                    logger.warning(f"Original filename is None or invalid for file_id: {file_id}")
            else:
                logger.warning(f"No file upload record found for file_id: {file_id}")
        except Exception as e:
            logger.warning(f"Could not determine original filename for file_id {file_id}: {e}")

        # Ensure buffer is at start and return a streaming response with correct headers
        excel_buffer.seek(0)
        from fastapi.responses import StreamingResponse
        response = StreamingResponse(
            content=excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        # Use Content-Disposition with quoted filename
        response.headers["Content-Disposition"] = f'attachment; filename="{original_filename}"'
        try:
            response.headers["Content-Length"] = str(len(excel_buffer.getvalue()))
        except Exception:
            pass
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download processed file error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Download error: {str(e)}"
        )

@app.get("/api/database/status")
async def database_status(session_id: str):
    """Check database connection status with detailed information"""
    try:
        # Verify session
        verify_session(session_id)
        
        # Get database configuration details
        from database_config.postgresql_config import PostgreSQLConfig
        db_config = PostgreSQLConfig()
        config_data = db_config.get_connection_params()
        
        # Check database connection
        try:
            # Get a direct database connection using psycopg2
            import psycopg2
            connection_params = db_config.get_connection_params()
            connection = psycopg2.connect(**connection_params)
            
            if connection:
                # Get database version info
                cursor = connection.cursor()
                cursor.execute("SELECT version();")
                db_version = cursor.fetchone()[0]
                cursor.close()
                connection.close()
                
                return {
                    "status": "connected",
                    "message": "Database connection successful",
                    "details": {
                        "host": config_data.get("host", "Unknown"),
                        "port": config_data.get("port", "Unknown"),
                        "database": config_data.get("database", "Unknown"),
                        "user": config_data.get("user", "Unknown"),
                        "version": db_version.split(" on ")[0] if " on " in db_version else db_version,
                        "connection_type": "PostgreSQL"
                    }
                }
            else:
                return {
                    "status": "disconnected",
                    "message": "Database connection failed",
                    "details": {
                        "host": config_data.get("host", "Unknown"),
                        "port": config_data.get("port", "Unknown"),
                        "database": config_data.get("database", "Unknown"),
                        "connection_type": "PostgreSQL"
                    }
                }
        except Exception as e:
            return {
                "status": "error",
                "message": f"Database error: {str(e)}",
                "details": {
                    "host": config_data.get("host", "Unknown"),
                    "port": config_data.get("port", "Unknown"),
                    "database": config_data.get("database", "Unknown"),
                    "connection_type": "PostgreSQL",
                    "error": str(e)
                }
            }
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database status error: {str(e)}"
        )

@app.get("/api/files/list")
async def list_uploaded_files(session_id: str):
    """List files uploaded to the database"""
    try:
        # Verify session
        verify_session(session_id)
        
        # Get database connection
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2
        
        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        
        cursor = connection.cursor()
        cursor.execute("""
            SELECT id, file_name, upload_date, uploaded_by, processing_status, 
                   records_count, file_size, processing_error
            FROM file_upload 
            ORDER BY upload_date DESC 
            LIMIT 50
        """)
        
        files = cursor.fetchall()
        cursor.close()
        connection.close()
        
        file_list = []
        for file_row in files:
            file_list.append({
                "id": file_row[0],
                "file_name": file_row[1],
                "upload_date": file_row[2].isoformat() if file_row[2] else None,
                "uploaded_by": file_row[3],
                "processing_status": file_row[4],
                "records_count": file_row[5],
                "file_size": file_row[6],
                "processing_error": file_row[7]
            })
        
        return {
            "success": True,
            "files": file_list,
            "total_files": len(file_list)
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error retrieving files: {str(e)}"
        )

@app.get("/api/debug/jobs")
async def debug_jobs(session_id: str):
    """Debug endpoint to check processing jobs"""
    verify_session(session_id)
    return {
        "total_jobs": len(processing_jobs),
        "job_ids": list(processing_jobs.keys()),
        "jobs": {k: {
            "status": v.get("status"),
            "progress": v.get("progress"),
            "message": v.get("message"),
            "filename": v.get("filename")
        } for k, v in processing_jobs.items()}
    }

@app.get("/", response_class=HTMLResponse)
async def read_root():
    """Serve the React app"""
    static_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend", "build")
    index_file = os.path.join(static_dir, "index.html")
    if os.path.exists(index_file):
        with open(index_file, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read(), status_code=200)
    else:
        return HTMLResponse(content="<h1>Company Data Scraper API</h1><p>React frontend not built. Run 'npm run build' in the frontend directory.</p>", status_code=200)

def validate_file_headers(file_content: bytes, filename: str) -> Dict[str, Any]:
    """Validate that uploaded file has correct headers matching our template"""
    try:
        import io
        
        # Expected headers (standardized format - preferred)
        required_headers = ['Company Name', 'LinkedIn_URL']
        optional_headers = ['Website_URL', 'Company_Size', 'Industry', 'Revenue']
        all_expected_headers = required_headers + optional_headers
        
        # Alternative naming conventions (for backward compatibility)
        header_alternatives = {
            'Company Name': ['Company_Name', 'company_name'],
            'LinkedIn_URL': ['Company Linkedin', 'linkedin_url', 'Company_Linkedin'],
            'Website_URL': ['Website', 'Company_Website', 'company_website'],
            'Company_Size': ['Size', 'company_size'],
            'Industry': ['industry', 'Company_Industry'],
            'Revenue': ['company_revenue', 'Company_Revenue']
        }
        
        # Read file headers
        if filename.lower().endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(file_content))
        elif filename.lower().endswith('.csv'):
            df = pd.read_csv(io.BytesIO(file_content))
        else:
            return {"valid": False, "error": "Unsupported file format"}
        
        file_headers = list(df.columns)
        
        # Check for required headers
        missing_required = []
        found_headers = {}
        
        for required in required_headers:
            found = False
            if required in file_headers:
                found_headers[required] = required
                found = True
            else:
                # Check alternatives
                alternatives = header_alternatives.get(required, [])
                for alt in alternatives:
                    if alt in file_headers:
                        found_headers[required] = alt
                        found = True
                        break
            
            if not found:
                missing_required.append(required)
        
        # Check for optional headers
        found_optional = []
        for optional in optional_headers:
            if optional in file_headers:
                found_optional.append(optional)
            else:
                alternatives = header_alternatives.get(optional, [])
                for alt in alternatives:
                    if alt in file_headers:
                        found_optional.append(optional)
                        break
        
        # Check for unexpected headers
        recognized_headers = set()
        for expected in all_expected_headers:
            if expected in file_headers:
                recognized_headers.add(expected)
            alternatives = header_alternatives.get(expected, [])
            for alt in alternatives:
                if alt in file_headers:
                    recognized_headers.add(alt)
        
        unexpected_headers = [h for h in file_headers if h not in recognized_headers]
        
        if missing_required:
            return {
                "valid": False,
                "error": f"Missing required columns: {', '.join(missing_required)}. For best results, please download and use our standardized Excel template which includes the exact column names expected by the system.",
                "missing_required": missing_required,
                "found_headers": file_headers,
                "expected_headers": all_expected_headers,
                "recommendation": "Download the standardized template for optimal compatibility"
            }
        
        return {
            "valid": True,
            "found_required": list(found_headers.keys()),
            "found_optional": found_optional,
            "unexpected_headers": unexpected_headers,
            "total_rows": len(df),
            "header_mapping": found_headers
        }
        
    except Exception as e:
        return {"valid": False, "error": f"Error validating file: {str(e)}"}

@app.post("/api/files/validate-headers")
async def validate_file_headers_endpoint(file: UploadFile = File(...), session_id: str = ""):
    """Validate file headers without uploading - for immediate user feedback"""
    try:
        # Verify session
        verify_session(session_id)
        
        # Validate file type
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file selected")
        
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            raise HTTPException(status_code=400, detail="Only Excel and CSV files are supported")
        
        # Read file content for validation
        file_content = await file.read()
        
        # Validate headers
        validation_result = validate_file_headers(file_content, file.filename)
        
        return {
            "success": True,
            "validation": validation_result,
            "filename": file.filename
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Validation error: {str(e)}"
        )

@app.post("/api/files/upload-json")
async def upload_file_as_json(file: UploadFile = File(...), session_id: str = ""):
    """Upload file as JSON to file_upload table with enhanced session and user management"""
    try:
        # Verify session with enhanced validation
        session_data = verify_session(session_id)
        if not session_data:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid or expired session. Please login again."
            )
        
        # Get user info from session
        user_info = session_data.get('user_info', {})
        username = user_info.get('username', 'Web_User')
        user_id = user_info.get('id', 0)
        
        # Validate file
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file selected")
        
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            raise HTTPException(status_code=400, detail="Only Excel and CSV files are supported")
        
        # Read file content for validation
        file_content = await file.read()
        
        # Validate headers before processing
        validation_result = validate_file_headers(file_content, file.filename)
        
        if not validation_result["valid"]:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid file format: {validation_result['error']}"
            )
        
        # Initialize file processor with user context
        try:
            file_processor = FileUploadProcessor()
            
            # Create temporary file for processing
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as temp_file:
                temp_file.write(file_content)
                temp_file_path = temp_file.name
            
            try:
                # Upload as JSON with user-specific context
                file_upload_id = file_processor.upload_file_as_json(
                    temp_file_path, 
                    uploaded_by=username,
                    original_filename=file.filename,
                    user_id=user_id
                )
                
                if file_upload_id:
                    logger.info(f"✅ File uploaded as JSON: {file.filename} by user {username} (ID: {file_upload_id})")
                    response_payload = {
                        "success": True,
                        "file_upload_id": file_upload_id,
                        "message": f"File uploaded as JSON successfully (ID: {file_upload_id})",
                        "filename": file.filename,
                        "status": "pending_processing",
                        "uploaded_by": username
                    }
                    # Ensure all values are JSON serializable (convert numpy types etc)
                    return _clean_for_json_serialization(response_payload)
                else:
                    raise HTTPException(status_code=500, detail="Failed to upload file as JSON")
                    
            finally:
                # Clean up temp file
                try:
                    os.unlink(temp_file_path)
                except:
                    pass
                    
        except ImportError as e:
            raise HTTPException(status_code=500, detail=f"File processor not available: {e}")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload failed for user {username if 'username' in locals() else 'unknown'}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@app.post("/api/files/process/{file_id}")
async def process_existing_file(file_id: str, session_id: str, scraping_enabled: bool = True, ai_analysis_enabled: bool = False):
    """Start processing for an already-uploaded file (by file_upload_id)"""
    try:
        # Verify session
        session = verify_session(session_id)
        if not session:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid session")

        # Load raw file content from database
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2

        db_config = PostgreSQLConfig()
        conn_params = db_config.get_connection_params()
        conn = psycopg2.connect(**conn_params)
        cursor = conn.cursor()
        # Note: DB column is `file_name` (not `filename`) - select it as filename
        cursor.execute("SELECT raw_data, file_name FROM file_upload WHERE id = %s", (file_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()

        if not row or not row[0]:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Uploaded file not found or no raw data available")

        raw_data = row[0]
        filename = row[1] or 'uploaded_file.xlsx'

        # Normalize raw_data into bytes that CompanyDataProcessor expects
        import io
        import json
        import base64

        excel_bytes = None
        try:
            # If bytes were stored directly (preferred), use them
            if isinstance(raw_data, (bytes, bytearray)):
                excel_bytes = bytes(raw_data)

            # If a dict was stored (e.g., {'columns':..., 'data': [...]}) reconstruct to Excel bytes
            elif isinstance(raw_data, dict):
                if 'data' in raw_data:
                    df_reconstructed = pd.DataFrame(raw_data['data'])
                else:
                    # try to create DataFrame directly from dict
                    df_reconstructed = pd.DataFrame(raw_data)
                buf = io.BytesIO()
                df_reconstructed.to_excel(buf, index=False)
                excel_bytes = buf.getvalue()
                buf.close()

            # If a string was stored, it could be JSON or base64 or raw CSV/text
            elif isinstance(raw_data, str):
                # Try JSON first
                try:
                    parsed = json.loads(raw_data)
                    if isinstance(parsed, dict) and 'data' in parsed:
                        df_reconstructed = pd.DataFrame(parsed['data'])
                        buf = io.BytesIO()
                        df_reconstructed.to_excel(buf, index=False)
                        excel_bytes = buf.getvalue()
                        buf.close()
                    else:
                        # If parsed to something else, attempt to create DataFrame
                        try:
                            df_reconstructed = pd.DataFrame(parsed)
                            buf = io.BytesIO()
                            df_reconstructed.to_excel(buf, index=False)
                            excel_bytes = buf.getvalue()
                            buf.close()
                        except Exception:
                            # Fallback to base64 decode
                            excel_bytes = base64.b64decode(raw_data)
                except Exception:
                    # Not JSON — try base64 decode
                    try:
                        excel_bytes = base64.b64decode(raw_data)
                    except Exception:
                        # As a last resort, treat string as raw bytes (e.g., CSV text)
                        excel_bytes = raw_data.encode('utf-8')

            else:
                # Unknown type — attempt to coerce
                try:
                    excel_bytes = bytes(raw_data)
                except Exception:
                    excel_bytes = None
        except Exception as _ex:
            excel_bytes = None

        if not excel_bytes:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unable to reconstruct uploaded file bytes from database raw_data")

        # Create in-memory job with normalized bytes
        job_id = str(uuid.uuid4())
        processing_jobs[file_id] = {
            "id": job_id,
            "file_id": file_id,
            "filename": filename,
            "status": "processing",
            "progress": 0,
            "message": "Processing started",
            "file_content": excel_bytes,
            "result": None
        }

        # Immediately update DB status to 'processing' so UI reflects the change
        try:
            from database_config.file_upload_processor import FileUploadProcessor
            fup = FileUploadProcessor()
            try:
                fup.update_processing_status(file_id, 'processing')
                print(f"✅ Database status updated to 'processing' for file: {file_id}")
            except Exception as upd_err:
                print(f"⚠️ Failed to update DB status to 'processing' for file {file_id}: {upd_err}")
        except Exception as import_err:
            print(f"⚠️ Could not import FileUploadProcessor to update DB status: {import_err}")

        # Update per-user in-memory processing status (best-effort)
        try:
            username = session.get('user_info', {}).get('username') if session else None
            if username:
                user_processing_status[username] = {
                    'status': 'processing',
                    'file_id': file_id,
                    'filename': filename,
                    'started': datetime.now().isoformat()
                }
        except Exception:
            pass

        # Start processing in background thread
        def _start():
            try:
                sys.path.insert(0, os.path.dirname(__file__))
                from company_processor import CompanyDataProcessor
                processor = CompanyDataProcessor()

                def update_progress(percent, message):
                    processing_jobs[file_id]["progress"] = percent
                    processing_jobs[file_id]["message"] = message

                # Use the normalized bytes stored in the job (ensures bytes-like object)
                job_bytes = processing_jobs[file_id].get("file_content")
                result = processor.process_file(
                    file_content=job_bytes,
                    filename=filename,
                    scraping_enabled=scraping_enabled,
                    ai_analysis_enabled=ai_analysis_enabled,
                    progress_callback=update_progress
                )

                processing_jobs[file_id]["result"] = result
                processing_jobs[file_id]["progress"] = 100
                processing_jobs[file_id]["status"] = "completed" if result.get("success") else "failed"
                processing_jobs[file_id]["message"] = result.get("summary", "Processing finished")

                # Sync status to database using FileUploadProcessor helper
                try:
                    from database_config.file_upload_processor import FileUploadProcessor
                    file_processor = FileUploadProcessor()
                    processed_count = int(result.get('processed_rows', 0)) if result.get('processed_rows') is not None else 0
                    if result.get('success'):
                        file_processor.sync_processing_completion(file_id, 'completed', processed_count)
                    else:
                        err = result.get('error', None) or result.get('traceback', '')
                        file_processor.sync_processing_completion(file_id, 'failed', 0, str(err))
                except Exception as dbsync_err:
                    # Log but don't raise — keep job status updated
                    print(f"❌ Failed to sync processing completion to DB for file {file_id}: {dbsync_err}")
            except Exception as e:
                processing_jobs[file_id]["status"] = "failed"
                processing_jobs[file_id]["message"] = str(e)

        _threading.Thread(target=_start, daemon=True).start()

        return {"success": True, "message": "Processing started", "file_id": file_id}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error starting processing for {file_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error starting processing: {str(e)}")

@app.post("/api/files/upload-and-process")
async def upload_and_process_file(file: UploadFile = File(...), session_id: str = ""):
    """Upload file as JSON and immediately process it with concurrent user support"""
    try:
        # Verify session with enhanced validation
        session_data = verify_session(session_id)
        if not session_data:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid or expired session. Please login again."
            )
        try:
            # Verify session with enhanced validation
            session_data = verify_session(session_id)
            if not session_data:
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="Invalid or expired session. Please login again."
                )
            # Get user info from session
            user_info = session_data.get('user_info', {})
            username = user_info.get('username', 'Web_User')
            user_id = user_info.get('id', 0)
            # Validate file
            if not file.filename:
                raise HTTPException(status_code=400, detail="No file selected")
            if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
                raise HTTPException(status_code=400, detail="Only Excel and CSV files are supported")
            file_content = await file.read()
            validation_result = validate_file_headers(file_content, file.filename)
            if not validation_result["valid"]:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid file format: {validation_result['error']}"
                )
            file_processor = None
            temp_file_path = None
            try:
                file_processor = FileUploadProcessor()
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as temp_file:
                    temp_file.write(file_content)
                    temp_file_path = temp_file.name
                file_upload_id = file_processor.upload_file_as_json(
                    temp_file_path,
                    uploaded_by=username,
                    original_filename=file.filename
                )
                if file_upload_id:
                    # Create a processing job (status: pending)
                    file_processor.create_processing_job(file_upload_id, job_type="data_extraction", uploaded_by=username)
                    response = {
                        "success": True,
                        "file_upload_id": file_upload_id,
                        "message": f"File uploaded successfully (ID: {file_upload_id})",
                        "filename": file.filename,
                        "status": "pending",
                        "processed_by": username
                    }
                    from backend_api.main import _clean_for_json_serialization
                    return _clean_for_json_serialization(response)
                else:
                    raise HTTPException(status_code=500, detail="Failed to upload file as JSON")
            except Exception as e:
                logger.error(f"❌ File processor error: {str(e)}")
                raise HTTPException(status_code=500, detail=f"File processor error: {str(e)}")
            finally:
                # Clean up temp file if created
                if temp_file_path:
                    try:
                        # 'os' is already imported globally
                        os.unlink(temp_file_path)
                    except Exception:
                        pass
        except Exception as e:
            logger.error(f"❌ Error in upload_and_process_file: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Upload and process error: {str(e)}")
        session_data = verify_session(session_id)
        if not session_data:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid or expired session. Please login again."
            )
        
        # Get user info from session
        user_info = session_data.get('user_info', {})
        username = user_info.get('username', 'Web_User')
        
        # Get current processing status for this user
        async with job_queue_lock:
            current_status = user_processing_status.get(username, {
                'status': 'idle',
                'message': 'No active processing jobs'
            })
        
        return {
            "success": True,
            "username": username,
            "processing_status": current_status,
            "queue_info": {
                "total_active_jobs": len([
                    status for status in user_processing_status.values() 
                    if status.get('status') == 'processing'
                ]),
                "user_has_active_job": current_status.get('status') == 'processing'
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get processing status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to get processing status: {str(e)}")

@app.get("/api/files/uploads")
async def get_uploaded_files(session_id: str):
    """Get list of uploaded files from file_upload table with enhanced session validation"""
    import time
    start_time = time.time()
    try:
        # Verify session with enhanced validation
        session_data = verify_session(session_id)
        if not session_data:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid or expired session. Please login again."
            )

        # Query real uploaded files from database and adapt to UI's expected shape
        try:
            from database_config.postgresql_config import PostgreSQLConfig
            import psycopg2

            db_config = PostgreSQLConfig()
            connection_params = db_config.get_connection_params()
            connection = psycopg2.connect(**connection_params)

            cursor = connection.cursor()
            cursor.execute(
                """
                SELECT fu.id, fu.file_name, fu.upload_date, fu.uploaded_by, fu.processing_status, 
                       fu.records_count,
                       COUNT(cd.id) as total_records,
                       COUNT(CASE WHEN cd.processing_status = 'completed' THEN 1 END) as processed_count,
                       COUNT(CASE WHEN cd.processing_status = 'failed' THEN 1 END) as failed_count
                FROM file_upload fu
                LEFT JOIN company_data cd ON fu.id = cd.file_upload_id
                GROUP BY fu.id, fu.file_name, fu.upload_date, fu.uploaded_by, fu.processing_status, fu.records_count
                ORDER BY fu.upload_date DESC 
                LIMIT 10
                """
            )

            rows = cursor.fetchall()
            cursor.close()
            connection.close()

            files = []
            for r in rows:
                file_info = {
                    "id": r[0],
                    "filename": r[1],
                    "file_name": r[1],  # Add both variants for compatibility
                    "upload_date": r[2].isoformat() if r[2] else None,
                    "uploaded_by": r[3],
                    "status": r[4],
                    "processing_status": r[4],
                    "total_rows": r[5],
                    "records_count": r[6],  # Actual count from company_data table
                    "processed_count": r[7],
                    "failed_count": r[8]
                }
                # Clean all values for JSON serialization
                from backend_api.main import _clean_for_json_serialization
                files.append(_clean_for_json_serialization(file_info))

            elapsed = time.time() - start_time
            return {"files": files, "count": len(files)}

        except ImportError as e:
            return {"files": [], "count": 0, "error": str(e)}
        except Exception as e:
            return {"files": [], "count": 0, "error": str(e)}
    except HTTPException as e:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get uploaded files: {str(e)}")

@app.delete("/api/files/{file_id}")
async def delete_file_and_data(file_id: str, session_id: str = ""):
    """Delete file and all associated data from all tables (file_upload, processing_jobs, company_data)"""
    try:
        # Verify session
        verify_session(session_id)
        
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2
        
        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        connection.autocommit = False  # Use transaction
        
        try:
            cursor = connection.cursor()
            
            # First, check if the file exists and get its info
            cursor.execute(
                "SELECT file_name, uploaded_by FROM file_upload WHERE id = %s",
                (file_id,)
            )
            file_info = cursor.fetchone()
            
            if not file_info:
                raise HTTPException(status_code=404, detail="File not found")
            
            file_name, uploaded_by = file_info
            logger.info(f"🗑️ Deleting file and data: {file_name} (ID: {file_id})")
            
            # Delete from company_data table (linked by file_upload_id)
            cursor.execute(
                "DELETE FROM company_data WHERE file_upload_id = %s",
                (file_id,)
            )
            company_data_deleted = cursor.rowcount
            logger.info(f"🗑️ Deleted {company_data_deleted} records from company_data")
            
            # Delete from processing_jobs table (linked by file_upload_id)
            cursor.execute(
                "DELETE FROM processing_jobs WHERE file_upload_id = %s",
                (file_id,)
            )
            processing_jobs_deleted = cursor.rowcount
            logger.info(f"🗑️ Deleted {processing_jobs_deleted} records from processing_jobs")
            
            # Finally delete from file_upload table
            cursor.execute(
                "DELETE FROM file_upload WHERE id = %s",
                (file_id,)
            )
            file_upload_deleted = cursor.rowcount
            
            if file_upload_deleted == 0:
                connection.rollback()
                raise HTTPException(status_code=404, detail="File not found")
            
            # Commit the transaction
            connection.commit()
            logger.info(f"✅ Successfully deleted file: {file_name} and all associated data")
            
            cursor.close()
            connection.close()
            
            return {
                "success": True,
                "message": f"File '{file_name}' and all associated data deleted successfully",
                "deleted_counts": {
                    "file_upload": file_upload_deleted,
                    "processing_jobs": processing_jobs_deleted,
                    "company_data": company_data_deleted
                }
            }
            
        except Exception as e:
            connection.rollback()
            cursor.close()
            connection.close()
            raise e
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error deleting file {file_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete file: {str(e)}")


# -------------------- Data Management endpoints --------------------

@app.get("/api/files/view-data/{file_id}")
async def view_file_data(file_id: str, session_id: str, limit: int = 100, offset: int = 0):
    """View processed company data for a specific file with pagination"""
    try:
        # Verify session
        verify_session(session_id)
        
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2
        
        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        cursor = connection.cursor()
        
        # Get file info first
        cursor.execute(
            """SELECT file_name, uploaded_by, upload_date, processing_status 
               FROM file_upload WHERE id = %s""",
            (file_id,)
        )
        file_info = cursor.fetchone()
        
        if not file_info:
            cursor.close()
            connection.close()
            raise HTTPException(status_code=404, detail="File not found")
        
        # Get total count
        cursor.execute(
            "SELECT COUNT(*) FROM company_data WHERE file_upload_id = %s",
            (file_id,)
        )
        total_count = cursor.fetchone()[0]
        
        # Get paginated company data
        cursor.execute(
            """SELECT id, company_name, linkedin_url, company_website, 
                      company_size, industry, revenue, processing_status, 
                      upload_date, updated_at
               FROM company_data 
               WHERE file_upload_id = %s 
               ORDER BY upload_date DESC
               LIMIT %s OFFSET %s""",
            (file_id, limit, offset)
        )
        
        records = cursor.fetchall()
        company_data = []
        
        for record in records:
            company_data.append({
                "id": record[0],
                "company_name": record[1],
                "linkedin_url": record[2],
                "company_website": record[3],
                "company_size": record[4],
                "industry": record[5],
                "revenue": record[6],
                "processing_status": record[7] or "pending",
                "created_at": record[8],  # This is actually upload_date from DB
                "updated_at": record[9]
            })
        
        cursor.close()
        connection.close()
        
        return {
            "success": True,
            "data": company_data,
            "total_records": total_count,
            "file_info": {
                "file_name": file_info[0],
                "uploaded_by": file_info[1],
                "upload_date": file_info[2],
                "processing_status": file_info[3]
            },
            "pagination": {
                "limit": limit,
                "offset": offset,
                "total": total_count
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error viewing data for file {file_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error retrieving data: {str(e)}")

@app.put("/api/files/edit-data/{record_id}")
async def edit_company_record(record_id: str, session_id: str, update_data: dict):
    """Edit a specific company data record"""
    try:
        # Verify session
        verify_session(session_id)
        
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2
        from datetime import datetime
        
        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        cursor = connection.cursor()
        
        # Build dynamic update query
        update_fields = []
        update_values = []
        
        allowed_fields = ['company_name', 'linkedin_url', 'company_website', 'company_size', 'industry', 'revenue']
        
        for field, value in update_data.items():
            if field in allowed_fields and value is not None:
                update_fields.append(f"{field} = %s")
                update_values.append(value)
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="No valid fields to update")
        
        # Add updated_at timestamp
        update_fields.append("updated_at = %s")
        update_values.append(datetime.now())
        update_values.append(record_id)
        
        query = f"""
            UPDATE company_data 
            SET {', '.join(update_fields)}
            WHERE id = %s
            RETURNING id, company_name, linkedin_url, company_website, 
                     company_size, industry, revenue, processing_status, 
                     upload_date, updated_at
        """
        
        cursor.execute(query, update_values)
        updated_record = cursor.fetchone()
        
        if not updated_record:
            cursor.close()
            connection.close()
            raise HTTPException(status_code=404, detail="Company record not found")
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return {
            "success": True,
            "message": "Company record updated successfully",
            "data": {
                "id": updated_record[0],
                "company_name": updated_record[1],
                "linkedin_url": updated_record[2],
                "company_website": updated_record[3],
                "company_size": updated_record[4],
                "industry": updated_record[5],
                "revenue": updated_record[6],
                "processing_status": updated_record[7],
                "created_at": updated_record[8],  # This is actually upload_date from DB
                "updated_at": updated_record[9]
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating company record {record_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating record: {str(e)}")

@app.delete("/api/files/delete-record/{record_id}")
async def delete_company_record(record_id: str, session_id: str):
    """Delete a specific company data record"""
    try:
        # Verify session
        verify_session(session_id)
        
        from database_config.postgresql_config import PostgreSQLConfig
        import psycopg2
        
        db_config = PostgreSQLConfig()
        connection_params = db_config.get_connection_params()
        connection = psycopg2.connect(**connection_params)
        cursor = connection.cursor()
        
        cursor.execute(
            "DELETE FROM company_data WHERE id = %s RETURNING company_name",
            (record_id,)
        )
        
        deleted_record = cursor.fetchone()
        
        if not deleted_record:
            cursor.close()
            connection.close()
            raise HTTPException(status_code=404, detail="Company record not found")
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return {
            "success": True,
            "message": f"Company record '{deleted_record[0]}' deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting company record {record_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting record: {str(e)}")


# -------------------- Scheduler control endpoints --------------------
@app.post("/api/jobs/process-pending")
async def run_process_pending_now(session_id: str = ""):
    """Run processing of pending uploads immediately in background (one-shot)."""
    # Verify session (optional: could enforce admin)
    if session_id:
        try:
            verify_session(session_id)
        except Exception:
            # proceed even if session invalid; comment to enforce
            pass

    # Fire-and-forget thread to run the job once
    _threading.Thread(target=_process_pending_uploads, daemon=True).start()
    return {"success": True, "message": "Pending uploads processing started"}


@app.post("/api/jobs/scheduler/start")
async def start_scheduler(session_id: str = "", mode: str = "cron", interval_minutes: int = 2):
    """Start APScheduler job that processes pending uploads periodically.

    mode: 'cron' (default, runs every 2 minutes) or 'interval' (every N minutes)
    interval_minutes: used only for interval mode
    """
    if session_id:
        try:
            verify_session(session_id)
        except Exception:
            pass

    with _scheduler_lock:
        sched = _ensure_scheduler()

        # Remove existing job if present to avoid duplicates
        try:
            existing = sched.get_job(SCHEDULER_JOB_ID)
            if existing:
                sched.remove_job(SCHEDULER_JOB_ID)
        except Exception:
            pass

        if mode == "interval":
            trigger = IntervalTrigger(minutes=max(1, int(interval_minutes)))
            schedule_desc = f"every {max(1, int(interval_minutes))} minutes"
        else:
            # Default cron every 2 minutes
            scheduler_interval = get_scheduler_interval()
            trigger = CronTrigger(minute=f"*/{scheduler_interval}")
            schedule_desc = "every 2 minutes"

        sched.add_job(
            _process_pending_uploads,
            trigger=trigger,
            id=SCHEDULER_JOB_ID,
            max_instances=1,
            coalesce=True,
            misfire_grace_time=300,
            replace_existing=True,
        )
        scheduler_state["job_added"] = True

        next_run = sched.get_job(SCHEDULER_JOB_ID).next_run_time.isoformat() if sched.get_job(SCHEDULER_JOB_ID) and sched.get_job(SCHEDULER_JOB_ID).next_run_time else None

    return {
        "success": True,
        "message": f"Scheduler started ({schedule_desc})",
        "next_run": next_run,
    }


@app.post("/api/jobs/scheduler/interval")
async def set_scheduler_interval_endpoint(minutes: int, session_id: str = ""):
    """Set the scheduler interval (minutes) and reschedule the job if it's running.

    This will attempt to persist the new interval to the project's config.json. If persistence fails,
    the in-memory config will be used for the running process until restart.
    """
    if session_id:
        try:
            verify_session(session_id)
        except Exception:
            pass

    from database_config.config_loader import set_scheduler_interval, get_scheduler_interval

    try:
        minutes = int(minutes)
        if minutes < 1:
            raise HTTPException(status_code=400, detail="Interval must be >= 1 minute")
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid minutes parameter")

    ok = set_scheduler_interval(minutes)
    if not ok:
        # Still continue; config may be in-memory only
        logger.warning("Scheduler interval set in-memory but failed to persist to config file")

    # If the scheduler is running with a job, restart it with the new interval
    with _scheduler_lock:
        if scheduler and scheduler.running and scheduler.get_job(SCHEDULER_JOB_ID):
            try:
                # Remove existing job and restart with cron trigger using new interval
                scheduler.remove_job(SCHEDULER_JOB_ID)
                from apscheduler.triggers.cron import CronTrigger
                trigger = CronTrigger(minute=f"*/{minutes}")
                scheduler.add_job(
                    _process_pending_uploads,
                    trigger=trigger,
                    id=SCHEDULER_JOB_ID,
                    max_instances=1,
                    coalesce=True,
                    misfire_grace_time=300,
                    replace_existing=True,
                )
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Failed to reschedule job: {e}")

    return {"success": True, "minutes": minutes, "message": "Scheduler interval updated"}


@app.post("/api/jobs/scheduler/stop")
async def stop_scheduler(session_id: str = ""):
    """Stop the periodic scheduler job (does not stop a currently running run)."""
    if session_id:
        try:
            verify_session(session_id)
        except Exception:
            pass

    with _scheduler_lock:
        if scheduler and scheduler.running:
            try:
                if scheduler.get_job(SCHEDULER_JOB_ID):
                    scheduler.remove_job(SCHEDULER_JOB_ID)
                    scheduler_state["job_added"] = False
                # Keep scheduler alive; do not shutdown to avoid uvicorn reload issues
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Failed to stop scheduler: {e}")
        else:
            scheduler_state["job_added"] = False

    return {"success": True, "message": "Scheduler stopped"}


@app.get("/api/jobs/scheduler/status")
async def scheduler_status():
    """Return scheduler and job status for UI diagnostics."""
    job_info = None
    if scheduler and scheduler.running:
        try:
            job = scheduler.get_job(SCHEDULER_JOB_ID)
            if job:
                job_info = {
                    "id": job.id,
                    "next_run_time": job.next_run_time.isoformat() if job.next_run_time else None,
                }
        except Exception:
            job_info = None
    return {
        "scheduler_running": bool(scheduler and scheduler.running),
        "job_present": bool(job_info is not None),
        "job": job_info,
        "state": scheduler_state,
    }



@app.get("/api/sample-template")
async def download_sample_template():
    """Download sample Excel template for company data upload"""
    try:
        sample_file_path = os.path.join(parent_dir, "assets", "sample_company_data.xlsx")
        
        if not os.path.exists(sample_file_path):
            # Create sample file if it doesn't exist with standardized columns
            import pandas as pd
            sample_data = {
                'Company Name': [
                    'Acme Corporation',
                    'Tech Solutions Ltd', 
                    'Global Industries',
                    'StartupXYZ',
                    'Enterprise Corp'
                ],
                'LinkedIn_URL': [
                    'https://www.linkedin.com/company/acme-corp',
                    'https://www.linkedin.com/company/tech-solutions',
                    'https://www.linkedin.com/company/global-industries',
                    'https://www.linkedin.com/company/startupxyz',
                    'https://www.linkedin.com/company/enterprise-corp'
                ],
                'Website_URL': [
                    'https://www.acme.com',
                    'https://www.techsolutions.com',
                    'https://www.globalind.com',
                    'https://www.startupxyz.com',
                    'https://www.enterprise.com'
                ],
                'Company_Size': [
                    '51-200 employees',
                    '11-50 employees',
                    '201-500 employees', 
                    '1-10 employees',
                    '1001-5000 employees'
                ],
                'Industry': [
                    'Technology',
                    'Software Development',
                    'Manufacturing',
                    'Fintech',
                    'Consulting'
                ],
                'Revenue': [
                    '$5M',
                    '$2M',
                    '$50M',
                    '$500K', 
                    '$200M'
                ]
            }
            df = pd.DataFrame(sample_data)
            
            # Ensure assets directory exists
            assets_dir = os.path.join(parent_dir, "assets")
            os.makedirs(assets_dir, exist_ok=True)
            
            df.to_excel(sample_file_path, index=False, sheet_name='Company Data')
            logger.info("Sample template file created with standardized columns")
        
        return FileResponse(
            path=sample_file_path,
            filename="company_data_template.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        logger.error(f"Error creating sample template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create sample template: {str(e)}"
        )

@app.on_event("startup")
async def startup_event():
    """Initialize services on startup"""
    logger.info("🚀 Starting up application services...")
    
    # Auto-start scheduler with config settings
    try:
        interval_minutes = get_scheduler_interval()
        logger.info(f"📅 Auto-starting scheduler with {interval_minutes} minute interval")
        
        # Start the scheduler automatically
        _ensure_scheduler()
        if scheduler and not scheduler.running:
            scheduler.start()
            logger.info("✅ Scheduler started successfully")
        
        # Add the job with config-based interval
        with _scheduler_lock:
            global scheduler_state
            if APSCHEDULER_AVAILABLE and scheduler:
                try:
                    # Remove existing job if any
                    try:
                        scheduler.remove_job(SCHEDULER_JOB_ID)
                    except:
                        pass
                    
                    # Add job with interval from config
                    scheduler.add_job(
                        _process_pending_uploads,
                        IntervalTrigger(minutes=interval_minutes),
                        id=SCHEDULER_JOB_ID,
                        replace_existing=True,
                        max_instances=1
                    )
                    
                    scheduler_state["running"] = True
                    scheduler_state["job_added"] = True
                    logger.info(f"✅ Scheduled job added: process pending uploads every {interval_minutes} minutes")
                    
                except Exception as e:
                    logger.error(f"❌ Failed to add scheduled job: {e}")
                    scheduler_state["last_error"] = str(e)
            else:
                logger.warning("⚠️ APScheduler not available - jobs will need to be processed manually")
    except Exception as e:
        logger.error(f"❌ Failed to auto-start scheduler: {e}")

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "version": "1.0.0",
        "scheduler_running": scheduler_state.get("running", False) if APSCHEDULER_AVAILABLE else False
    }

if __name__ == "__main__":
    import uvicorn
    print("🚀 Starting FastAPI server on http://localhost:8000")
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)